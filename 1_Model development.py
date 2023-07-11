# Formulation of a ML-based model for the Assessment of Maximum Sprint Capability in Elite Soccer Players

# --- Import packages ---

import pandas as pd
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
import numpy as np
import plotly.express as px
from sklearn import preprocessing, linear_model, model_selection, metrics, cluster
import scipy
import statsmodels.api as sm
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.utils import get_column_letter
import matplotlib.cm as cm

from sklearn_extra.cluster import KMedoids
from sklearn.model_selection import ParameterGrid

import warnings

warnings.filterwarnings('ignore')

# ---


# --- Function  to compute the confidence interval ---
def ICpoly(xx_poly, X, ssr):

    N = X.shape[0]  # number of samples
    m = X.shape[1]  # number of features
    t = stats.t(df=(N-m)).ppf(0.975)  # compute the percentile
    rad = np.sqrt(
        np.sum((xx_poly @ (np.linalg.inv(X.T @ X))) * xx_poly, axis=1) + 1)

    return t * np.sqrt(ssr / (N-m)) * rad
# ---

# --- Function to evaluate the gaussianity of some data ---


def GaussianityTesting(data, test='AD', plotting=False):
    # Anderson-Darling test
    if test == 'AD':
        anderson_result = stats.anderson(data, dist='norm')
        critical_value = anderson_result.critical_values[2]
        statistic = anderson_result.statistic
        p_value = statistic < critical_value

    # Shapiro-Wilk test
    if test == 'SW':
        statistic, p_value = stats.shapiro(data)

    # Kolmogorov-Smirnov test
    if test == 'KS':
        statistic, p_value = stats.kstest(data, 'norm')

    # Pearson test
    if test == 'P':
        statistic, p_value = stats.normaltest(data)

    if plotting is True:
        sm.qqplot(data, line='45')
        plt.show()

    if p_value > 0.04999 or p_value is True:
        return True
    return False

# ---


# --- Function to evaluate the significance of the model's parameters ---
def SignificanceEvaluation(parameters, data, ssr):

    N = len(data)
    mse = ssr / (N - len(parameters))

    poly_reg = preprocessing.PolynomialFeatures(degree=len(parameters) - 1)
    x_poly = poly_reg.fit_transform(data.reshape(-1, 1))

    std_err = np.sqrt(np.diagonal(mse * np.linalg.inv(x_poly.T @ x_poly)))
    t_values = parameters / std_err
    p_values = (1 - stats.t.cdf(np.abs(t_values),
                df=(N - len(parameters)))) * 2

    ret = np.zeros(len(parameters), dtype=bool)
    ret = p_values < 0.05

    return ret
# ---

# --- Function to compute the value of Akaike Information Criterion ---


def AIC(degree, log_likelihood):

    return log_likelihood + 2 * (degree+1)
# ---

# --- Function to compute the value of Bayesian Information Criterion ---


def BIC(degree, log_likelihood, x):

    return log_likelihood + np.log(len(x)) * (degree+1)
# ---


# --- Function to compute the accelerative profile ---

def AS_profile(data, degree, n_peaks, step, threshold=99.99, plotting=False, results=False):

    x = data[:, -2].astype(np.float32)
    y = data[:, -1].astype(np.float32)

    not_none_idx = np.where(np.equal(y, None) == False)
    x = x[not_none_idx]
    y = y[not_none_idx]

    positive_idx = np.where(y > 0)[0]
    x = x[positive_idx]
    y = y[positive_idx]

    not_out_idx = np.where(y <= np.percentile(y, threshold))
    x = x[not_out_idx]
    y = y[not_out_idx]

    xx = np.array([])
    yy = np.array([])

    flag = True

    start = 0
    if degree == 1:
        start = 3

    for i in np.arange(start, np.max(x), step):
        idx = np.where((x >= i) & (x <= i + step))[0]
        yyy = y[idx]
        xxx = x[idx]

        sort = np.argsort(yyy)
        yy = np.append(yy, yyy[sort][-n_peaks:])
        xx = np.append(xx, xxx[sort][-n_peaks:])

    poly_reg = preprocessing.PolynomialFeatures(degree=degree)
    x_poly = poly_reg.fit_transform(xx.reshape(-1, 1))
    lin_reg = linear_model.LinearRegression(fit_intercept=flag)
    lin_reg.fit(x_poly, yy)
    y_pred_poly = lin_reg.predict(x_poly)
    coefficienti = lin_reg.coef_
    coefficienti[0] = lin_reg.intercept_

    if results == True:
        model = sm.OLS(yy, x_poly)
        results = model.fit()
        print(results.summary())

    residuals = yy - y_pred_poly
    gauss1 = GaussianityTesting(residuals)
    ssr = np.sum(residuals**2)
    sign1 = SignificanceEvaluation(coefficienti, xx, ssr)

    ala = ICpoly(x_poly, x_poly, ssr)

    ub = y_pred_poly + ala
    lb = y_pred_poly - ala

    idx_to_take = np.where((yy < ub) & (yy > lb))[0]
    xx = xx[idx_to_take]
    yy = yy[idx_to_take]

    x_poly = poly_reg.fit_transform(xx.reshape(-1, 1))
    lin_reg_2 = linear_model.LinearRegression(fit_intercept=flag)
    lin_reg_2.fit(x_poly, yy)
    y_pred_poly = lin_reg_2.predict(x_poly)
    coefficienti = lin_reg_2.coef_
    coefficienti[0] = lin_reg_2.intercept_

    residuals = yy - y_pred_poly
    gauss2 = GaussianityTesting(residuals)
    ssr = np.sum(residuals ** 2)
    sign2 = SignificanceEvaluation(coefficienti, xx, ssr)

    x_surrogate = np.linspace(0, np.max(xx), 101)

    halfwidth = ICpoly(poly_reg.fit_transform(
        x_surrogate.reshape(-1, 1)), x_poly, ssr)
    y_surrogate = poly_reg.fit_transform(
        x_surrogate.reshape(-1, 1)) @ coefficienti

    up = y_surrogate + halfwidth
    low = y_surrogate - halfwidth

    R2 = metrics.r2_score(yy, y_pred_poly)

    if plotting == True:
        plt.figure(figsize=(10, 6))
        sns.scatterplot(x=x, y=y, label='Starting samples')
        sns.scatterplot(x=xx, y=yy, color='red', label='Peaks', marker="X")
        sns.lineplot(x=x_surrogate, y=y_surrogate, color='blue',
                     label='Regression curve')
        plt.fill_between(x_surrogate, up, low, color='gray',
                         alpha=0.25, label='Confidence interval to 95%')

        plt.xlabel('Speed')
        plt.ylabel('Acceleration')
        plt.title(f'Polynomial regression with order {degree}')
        plt.legend(fontsize='small', framealpha=1, shadow=True)
        plt.show()

    if results == True:
        model = sm.OLS(yy, x_poly)
        results = model.fit()
        print(results.summary())

    IC = [ICpoly(poly_reg.fit_transform(np.array(0).reshape(-1, 1)), x_poly, ssr), ICpoly(poly_reg.fit_transform(
        np.max(xx).reshape(-1, 1)), x_poly, ssr), ICpoly(poly_reg.fit_transform(np.mean(xx).reshape(-1, 1)), x_poly, ssr)]

    residuals = yy - y_pred_poly

    log_likelihood = len(xx) * (np.log(2 * np.pi) + np.log(ssr/(len(xx))) + 1)

    R2adj = 1 - (1-R2)*(len(xx)-1)/(len(xx)-len(coefficienti))

    return R2, coefficienti,  [np.min(xx), np.max(xx)], np.max(y_pred_poly), \
        IC, ssr, [gauss1, gauss2, sign1, sign2, residuals], [
            AIC(log_likelihood, degree), BIC(log_likelihood, degree, xx), R2adj], xx, yy
# ---


# --- STAMINA EVALUATION ---
def stamina(all_players, blocks=4):

    time_block = 90/blocks
    intervals = np.arange(time_block, 90, time_block) * 60

    output = [[] for _ in np.arange(len(intervals) + 1)]

    for player in all_players:

        movements = np.array(all_players[player]['tracking'])
        not_none_idx = np.where(np.equal(movements[:, -1], None) == False)
        movements = movements[not_none_idx]

        for i, threshold in enumerate(intervals):

            mask = np.less_equal(movements[:, 0], threshold)
            output[i].extend(movements[:, 1:][mask])
            movements = movements[~mask]

        output[-1].extend(movements[:, 1:])

    stamina_ary = np.array([])
    coeff = np.array([])

    for i in range(len(output)):
        stamina_ary = np.append(stamina_ary, AS_profile(
            np.array(output[i]), 2, 2, step=0.15)[3])

        coeff = np.append(coeff, AS_profile(
            np.array(output[i]), 2, 2, step=0.15)[1])

    stamina_ary = np.concatenate(
        (stamina_ary.reshape(-1, 1), stamina_ary.reshape(-1, 1) / stamina_ary[0]), axis=1)

    pd.DataFrame(stamina_ary, index=['22:30', '45:00', '67:30', '90:00'], columns=[
                 'Stamina', 'Percentual Variation'])

    return stamina_ary



# --- Stamina trend ---
def stamina_trend(intervals, stamina_ary):

    x = np.concatenate(np.array([0]), intervals)
    y = np.concatenate(np.array([0]), stamina_ary)
    df_i = pd.DataFrame({'x': x, 'y': y})

    coefficients_stamina = np.polyfit(x, y, 2)
    polynomial = np.poly1d(coefficients_stamina)

    x_curve = np.linspace(0, 90*60, 100)
    y_curve = polynomial(x_curve)

    fig = px.scatter(df_i, x='x', y='y')

    data = {'x': x + x_curve.tolist(), 'y': y + y_curve.tolist()}

    fig.add_trace(px.line(x=x_curve, y=y_curve).data[0])

    fig.update_layout(
        title='Stamina',
        xaxis_title='Time (seconds)',
        yaxis_title='Maximum acceleration capability'
    )

    fig.show()

    return coefficients_stamina
# ---




# --- SPRINT SIMULATION ---

def sprint_test(coefficients, coefficients_stamina, minute=0, plotting=False):

    model = coefficients[::-1]

    distance_list = [10, 20, 35, 50, 75, 100]
    v0_list = [0, 2.5, 5.5, 7]
    # sprint_list = []

    stamina = np.polyval(coefficients_stamina, minute*60)

    time_list = np.array([])

    for distance in distance_list:

        for v0 in v0_list:

            runned_distance = 0
            speed = v0
            time = 0

            posizione = np.array([])
            tempo = np.array([])
            velocit = np.array([])
            accs = np.array([])

            while runned_distance < distance:

                acceleration = np.polyval(model, speed) * stamina

                if acceleration > 0:

                    runned_distance += 0.5 * acceleration * 0.1 ** 2 + speed * 0.1
                    speed = speed + 0.1 * acceleration
                    time += 0.1

                else:
                    runned_distance += speed * 0.1
                    time += 0.1

                velocit = np.append(velocit, speed)
                posizione = np.append(posizione, runned_distance)
                tempo = np.append(tempo, time)
                accs = np.append(accs, acceleration)

            # sprint_list.append({'distance': distance, 'v0': v0, 'time': time})
            time_list = np.append(time_list, time)

            if plotting == True:

                data = {'Time': tempo, 'Speed': velocit}
                fig = px.line(data, x='Time', y='Speed')
                fig.update_layout(width=600, height=400)
                fig.show()

                data = {'Time': tempo, 'Acceleration': accs}
                fig = px.line(data, x='Time', y='Acceleration')
                fig.update_layout(width=600, height=400)
                fig.show()

                data = {'Time': tempo, 'Position': posizione}
                fig = px.line(data, x='Time', y='Position')
                fig.update_layout(width=600, height=400)
                fig.show()

    return time_list

# ---


def create_excel(labels, data, names):
    ord = np.argsort(labels)
    labels = labels[ord]
    data = data[ord]
    names = np.array(names)[ord]
    data = np.column_stack((data, labels))

    df = pd.DataFrame(data.T, columns=names)
    num_clusters = df.iloc[-1, :].nunique()
    colors = cm.tab20c(np.linspace(0, 1, num_clusters))

    wb = Workbook()
    ws = wb.active

    for i, name in enumerate(names):
        ws.cell(row=1, column=i+2, value=name)
        for j, val in enumerate(data.T[:-1, i]):
            ws.cell(row=j+2, column=i+2, value=val)

    def rgb_array_to_hex(rgb_array):
        hex_colors = []
        for color in rgb_array:
            r, g, b, _ = color
            hex_color = f"FF{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"
            hex_colors.append(hex_color)
        return hex_colors

    hex_colors = rgb_array_to_hex(colors)

    cluster_colors = {c: hex_colors[i] for i, c in enumerate(
        df.iloc[-1, :].unique()) if i < len(hex_colors)}

    for i, cluster in enumerate(labels):
        color = cluster_colors.get(cluster, Color(rgb="FFFFFF"))
        fill = PatternFill(patternType='solid', fgColor=color)
        column_letter = get_column_letter(i + 2)
        for cell in ws[column_letter]:
            cell.fill = fill

    wb.save('cluster_data.xlsx')

# ---

# --- Clustering ---
def clustering_method(data_dict, k=2):

    def custom_scorer(estimator, X):
        labels = estimator.labels_
        silhouette = metrics.silhouette_score(X, labels)
        calinski_harabasz = metrics.calinski_harabasz_score(X, labels)
        davies_bouldin = metrics.davies_bouldin_score(X, labels)
        return silhouette + calinski_harabasz/100 - davies_bouldin

    names = np.array(list(data_dict.keys()))
    data = np.array(list(data_dict.values()))[:, :]

    dict_test = {}

    # --- K-Means Algorithm ---

    kmeans = cluster.KMeans(n_clusters=k)

    parameters_kmeans = {
        'init': ['k-means++', 'random'],
        'max_iter': [100, 300, 500, 700, 1000]
    }

    grid_search_kmeans = model_selection.GridSearchCV(
        kmeans, parameters_kmeans, scoring=custom_scorer)
    grid_search_kmeans.fit(data)

    best_init = grid_search_kmeans.best_params_['init']
    best_max_iter = grid_search_kmeans.best_params_['max_iter']

    y_km = cluster.KMeans(n_clusters=k, init=best_init,
                          max_iter=best_max_iter).fit_predict(data)

    dict_test['K-means method'] = {'S': metrics.silhouette_score(data, y_km),
                                   'DB': metrics.davies_bouldin_score(data, y_km),
                                   'CH': metrics.calinski_harabasz_score(data, y_km)}

    # ---

    # Definizione dei parametri da testare
    param_grid = {
        'n_clusters': [2],
        'init': ['random', 'heuristic', 'k-medoids++', 'build'],
        'max_iter': [100, 300, 500, 700, 1000],
        'random_state': [1, 2, 3, 4, 5, 6],
        'method': ['pam']
    }

    # --- K-Medoids Algorithm ---
    best_score = -10
    best_params = {}

    for params in ParameterGrid(param_grid):

        kmedoids = KMedoids(**params)

        kmedoids.fit(data)

        score = custom_scorer(kmedoids, data)

        if score > best_score:
            best_score = score
            best_params = params

    kmedoids = KMedoids(**best_params)
    y_medoid = kmedoids.fit_predict(data)

    dict_test['K-Medoids method'] = {'S': metrics.silhouette_score(data, y_medoid),
                                     'DB': metrics.davies_bouldin_score(data, y_medoid),
                                     'CH': metrics.calinski_harabasz_score(data, y_medoid)}
    # ---

    # --- Hierarchical Agglomerative Algorithm ---

    link_matrix = scipy.cluster.hierarchy.linkage(data, method='ward')
    scipy.cluster.hierarchy.dendrogram(
        link_matrix, labels=names, orientation='left', color_threshold=2, leaf_rotation=0, count_sort='descending')

    y_hc = scipy.cluster.hierarchy.fcluster(
        link_matrix, t=k, criterion='maxclust')

    dict_test['Hierarchical clustering'] = {'S': metrics.silhouette_score(data, y_hc),
                                            'DB': metrics.davies_bouldin_score(data, y_hc),
                                            'CH': metrics.calinski_harabasz_score(data, y_hc)}

    # ---

    # --- BIRCH Algorithm ---

    birch = cluster.Birch(n_clusters=k)

    parameters = {
        'threshold': [0.1, 0.3, 0.5, 0.7],
        'branching_factor': [20, 50, 75]
    }

    grid_search_birch = model_selection.GridSearchCV(
        birch, parameters, scoring=custom_scorer)
    grid_search_birch.fit(data)

    best_threshold = grid_search_birch.best_params_['threshold']
    best_branching_factor = grid_search_birch.best_params_['branching_factor']

    model = cluster.Birch(branching_factor=k, n_clusters=cluster.AgglomerativeClustering(
        n_clusters=2), threshold=best_threshold)
    y_birch = model.fit_predict(data)

    dict_test['Birch algorithm'] = {'S': metrics.silhouette_score(data, y_birch),
                                    'DB': metrics.davies_bouldin_score(data, y_birch),
                                    'CH': metrics.calinski_harabasz_score(data, y_birch)}

    print(best_branching_factor, best_threshold)

    # ---

    cluster_dict = {i: [] for i in np.unique(y_birch)}
    for i, label in enumerate(y_birch):
        cluster_dict[label].append(names[i])

    for label, entity_names in cluster_dict.items():
        print(f'Cluster {label}: {entity_names}')

    result = np.array(pd.DataFrame(dict_test))

    for i in range(len(result)):
        maximum = np.max(result[i])
        for j in range(len(result[i])):
            result[i, j] = result[i, j] / maximum

    total = np.array([])
    for column in result.T:
        total = np.append(
            total, 1/3 * column[0] - 1/3 * column[1] + 1/3 * column[2])

    result = np.concatenate((result, total.reshape(1, 4)), axis=0)

    result = pd.DataFrame(result, index=['S', 'DB', 'CH', 'TS'],
                          columns=['K-Means Method', 'K-Medoids Method', 'Birch Algorithm', 'Hierarchical Clustering Algorithm'])

    create_excel(y_hc, data, names)

    return y_hc, pd.DataFrame(dict_test), result

# ---
